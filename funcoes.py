from openpyxl import load_workbook,Workbook
from datetime import date
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

try:
    arquivo = load_workbook("devedores.xlsx")    
except:
    arquivo = Workbook()
    
planilha1 = arquivo.active

def sobre():
    messagebox.showinfo(title="Debt Book",message="Programa gratuito desevolvido para distribuição com foco no pequeno comerciante, com este programa você será capaz de manusear dados sem ter nenhum conhecimento de planilhas.\nPara maiores informações ou sugestões: Rafaelmafra@live.com")

def cadastrar():
    limpa()
    clearFrame(quadro3)
    vdev=StringVar()
    lb1["text"] = "Cadastrar Cliente"
    dev = Entry(quadro,textvariable=vdev)
    dev.place(x=10,y=25,width=270,height=20)
    botao=Button(quadro, text="Verificar",command=lambda:cadastro(vdev.get()))
    botao.place(x=100,y=50,width=100,height=20)

def cadastronovo(valor, devedor,data):
    try:
        divida = str(valor)
        divida = divida.replace(",",".")
        divida = float(divida)
        dev = devedor
        atual = data
        if divida < 0:
            limpa()
            clearFrame(quadro3)
            lb2['text']= "Insira apenas valores positivos\nOperação cancelada!"

        elif divida > 0:
            devedores = divida,dev,atual,divida
            planilha1.append(devedores)
            save()
            limpa()
            clearFrame(quadro3)
            lb2['text']= "Valores Salvos\nOperação Concluída!"
            save()
    except:
        limpa()
        clearFrame(quadro3)
        lb2['text']= "Insira apenas valores Numericos\nOperação cancelada!"

def cadastro(user):
    dataatual = date.today()
    data = dataatual.strftime('%d/%m/%Y')
    try: 
        consultados = verifica_registro()
        devedor = str(user).lower()
        devedor = devedor.capitalize()
        
        if devedor not in consultados:
            vdiv = StringVar()
            tdiv = Entry(quadro3,textvariable=vdiv)
            tdiv.place(x=10,y=0,width=270,height=20)
            lb2['text'] = "Insira o valor da dívida"
            b1 = Button(quadro3, text="Registrar", command=lambda:cadastronovo(vdiv.get(),devedor,data))
            b1.place(x=50,y=50)
            
        else:
            max_coluna = planilha1.max_column
            lb2["text"]= "Usuario já cadastrado\nInforme o tipo de operação financeira:"
            b0 = Button(quadro3, text="Somar Dívida", command=lambda:entradasoma(max_coluna,data,devedor,consultados))
            b0.place(x=35,y=10)
            b1= Button(quadro3, text="Abater Divida", command=lambda:entradaabate(max_coluna,data,devedor,consultados))
            b1.place(x=170,y=10)
    
    except:
       
        print("Formato de entrada invalido")
        
def save():
    arquivo.save('devedores.xlsx')

def leitura():
    clearFrame(quadro2)
    lb1['text']="Lista De Devedores"
    max_linha= planilha1.max_row
    max_coluna= planilha1.max_column
    scroll_bar = Scrollbar(quadro)
    scroll_bar.pack( side = RIGHT, fill = Y)
    text = Text(quadro,bg="#5f9ea0",padx=10, yscrollcommand= scroll_bar.set)
    scroll_bar.config( command = text.yview ) 
    for i in range (1, max_linha+1):
        if planilha1.cell(row=i, column=1).value == "0.0" or planilha1.cell(row=i, column=1).value == None:
            continue
        text.insert(INSERT,".:"+planilha1.cell(row=i, column=2).value+":.\n")
        text.pack()
        for j in range(2, max_coluna+2):
            if planilha1.cell(row=i, column=j).value == None and planilha1.cell(row=i, column=j-1).value != None:
                text.insert(INSERT,planilha1.cell(row=i, column=j-2).value+" ")
                text.pack()
        text.insert(INSERT,"R$ {:.2f} \n\n".format(float(planilha1.cell(row=i, column=1).value)))
        text.pack()        
                          
def relatorio(): 
    lb1['text'] ="Relatório Detalhado"
    lb2['text']="Digite o nome do cliente a ser pesquisado"
    vuser = StringVar()   
    b1 = Entry(quadro3, textvariable=vuser)
    b1.place(x=85,y=0)
    b2 = Button(quadro3, text="Verificar",command=lambda:relatorioexec(vuser.get()) )
    b2.place(x=120,y=30)

def relatorioexec(user):
    max_coluna = planilha1.max_column
    clearFrame(quadro3)
    try:
        clearFrame(quadro2)
        quadro4 = Frame(quadro,bg="#5f9ea0")
        quadro4.place(x=0,y=78,width=285,height=265)
        lb1['text']="Relatório Detalhado"
        max_coluna= planilha1.max_column
        scroll_bar = Scrollbar(quadro4)
        scroll_bar.pack( side = RIGHT, fill = Y)
        text = Text(quadro4,bg="#5f9ea0",padx=10, yscrollcommand= scroll_bar.set)
        scroll_bar.config( command = text.yview ) 
        devedor = str(user).lower()
        devedor = devedor.capitalize()
        registro = verifica_registro()
        pos = registro.index(devedor)
        lb3["text"]="Exibindo resultados para || {} ||".format(devedor)
        for j in range(3, max_coluna+1):
            if planilha1.cell(row=pos+1, column=j).value != None:
                if j%2==0:
                    text.insert(INSERT,"R$: {:.2f}\n".format(float(planilha1.cell(row=pos+1, column=j).value)))
                    text.pack()
                else:
                    text.insert(INSERT,"{} - ".format(planilha1.cell(row=pos+1, column=j).value))
                    text.pack()
        total = planilha1.cell(row=pos+1, column=1).value
      
        text.insert(INSERT,"\nResultado Total: {:.2f}".format(float(total)))   
        
    except Exception as e: 
        print(repr(e))
        print("Usuario invalido")
    
    
def verifica_registro():
    consultados = []
    max_linha= planilha1.max_row
    for i in range(1,max_linha+1):
        consultado = planilha1.cell(row=i, column=2).value        
        consultados.append(consultado)
    return consultados

def excluir():
    limpa()
    clearFrame(quadro3)
    vdev=StringVar()
    lb1["text"] = "Deletar Cliente"
    dev = Entry(quadro,textvariable=vdev)
    dev.place(x=10,y=25,width=270,height=20)
    botao=Button(quadro, text="Deletar",command=lambda:excluiusuario(vdev.get()))
    botao.place(x=100,y=50,width=100,height=20)

def excluiusuario(user): 
    try: 
        consultados = verifica_registro()
        devedor = str(user).lower()
        devedor = devedor.capitalize()
        global pos 
        pos = consultados.index(devedor)
        lb2['text']= "Deseja realmente remover {}\nda lista? Você perdera os valores salvos".format(devedor)
        opcao1=Button(quadro3, text='Sim', command=opcao(1))
        opcao1.place(x=100, y=10)
        opcao2=Button(quadro3, text="Não", command=opcao(2))
        opcao2.place(x=150,y=10)
    except:
        lb2['text']='Ocorreu um erro, tente novamente!'
        
def limpa():
    lb2['text']= ""

def clearFrame(frame):
    # destroy all widgets from frame
    for widget in frame.winfo_children():
       widget.destroy()

def opsoma(informado,max_coluna,data,devedor,consultados):
    somar = str(informado)
    somar = somar.replace(",",".")
    somar = float(somar)  
    if somar < 0:
        lb2['text']="Insira apenas valores positivos\nOperação cancelada!"
        clearFrame(quadro3)
    elif somar > 0:
        pos = consultados.index(devedor)
        valor = float(planilha1['A'+str(pos+1)].value)+somar
        planilha1['A'+str(pos+1)] = str(valor)
        i = 5
        while i < max_coluna+2:
            if planilha1.cell(row=pos+1, column=i).value == None:
                planilha1.cell(row=pos+1, column=i).value = str(data)
                planilha1.cell(row=pos+1, column=i+1).value = str(somar)
                break
            i+=1
               
        lb2['text']="O valor R$ {:.2f} foi Somado com Sucesso!\nNovo Valor de R$ {:.2f} atualizado\npara o cliente {}".format(somar,float(planilha1['A'+str(pos+1)].value), devedor)
        clearFrame(quadro3)

def entradasoma(max_coluna,data,devedor,consultados):
    clearFrame(quadro3)
    lb1['text']="Somar Dívida"
    lb2["text"]="Insira O valor Para Somar"
    vsoma=StringVar()
    tsoma=Entry(quadro3, textvariable=vsoma)
    tsoma.place(x=10,y=10)
    b1 = Button(quadro3, text="Somar Valor", command=lambda:opsoma(vsoma.get(),max_coluna,data,devedor,consultados))
    b1.place(x=10,y=70)

def entradaabate(max_coluna,data,devedor,consultados):
    limpa()
    clearFrame(quadro3)
    lb1['text']="Abater Dívida"
    lb2["text"]="Insira O valor Para Abater"
    vabate=StringVar()
    tabate=Entry(quadro3, textvariable=vabate)
    tabate.place(x=10,y=10)
    b1 = Button(quadro3, text="Abater Valor", command=lambda:abater(vabate.get(),max_coluna,data,devedor,consultados))
    b1.place(x=10,y=70)
def abater(informado,max_coluna,data,devedor,consultados):   
    abate = str(informado)
    abate = abate.replace(",",".")
    abate = float(abate)
    if abate < 0:
        lb2['text']="Insira apenas valores positivos\nOperação cancelada!"
        clearFrame(quadro3)
    elif abate >0 :
        pos = consultados.index(devedor)
        valor = float(planilha1['A'+str(pos+1)].value)-abate
        planilha1['A'+str(pos+1)] = str(valor)
        i=5
        while i < max_coluna+2:
            if planilha1.cell(row=pos+1, column=i).value == None:
                planilha1.cell(row=pos+1, column=i).value = str(data)
                planilha1.cell(row=pos+1, column=i+1).value = "-"+str(abate)
                break
            i+=1
        lb2['text']="O valor R$ {:.2f} foi abatido com Sucesso!\nNovo Valor de R$ {:.2f} atualizado\npara o cliente {}".format(abate,float(planilha1['A'+str(pos+1)].value), devedor)
        clearFrame(quadro3)
       
def opcao(num):      
    if num == 1:
        planilha1.delete_rows(pos+1)
        arquivo.save('devedores.xlsx')
        clearFrame(quadro3)
        lb2['text'] = "Cliente Excluído da base de dados!!"
    elif num ==2:
        lb2['text']='Operação Cancelada!'
        clearFrame(quadro3)

def sair():
    arquivo.save('devedores.xlsx')
    exit()

def novo():
    quadro = Frame(app,borderwidth="2",bg="#5f9ea0",relief="groove")
    quadro.place(x=215,y=100,width=290,height=350)
    quadro2 = Frame(quadro,bg="#5f9ea0")
    quadro2.place(x=0,y=78,width=285,height=265)
    quadro3 = Frame(quadro2,bg="#5f9ea0")
    quadro3.place(x=0, y= 50,width=285,height=280)

app = Tk()

app.title(".::.Debt Book.::. Caderneta de Dívidas")
app.geometry("510x500+300+100")
app.configure(background="#5f9ea0")
quadro = Frame(app,borderwidth="2",bg="#5f9ea0",relief="groove")
quadro.place(x=215,y=100,width=290,height=350)
quadro2 = Frame(quadro,bg="#5f9ea0")
quadro2.place(x=0,y=78,width=285,height=265)
quadro3 = Frame(quadro2,bg="#5f9ea0")
quadro3.place(x=0, y= 50,width=285,height=280)

txt1 = Label(app,text=".::.Bem Vindo Ao Debt Book.::.",bg="#fff", fg="#000",font=('arial black',12),justify="center")
txt1.place(x=5, y=5, width=500,height=30)
txt2 = Label(app, text='Desenvolvido por Rafael Mafra, programa gratuito, proibido a venda',bg="#5f9ea0",fg="#fff",justify="center")
txt2.place(x=5, y=450, width=500, height=30)

barrademenus=Menu(app)
app.config(menu=barrademenus)
menusobre=Menu(barrademenus, tearoff=0)
menusobre.add_command(label="Info",command=sobre)
menusobre.add_command(label="Sair",command=app.quit)
barrademenus.add_cascade(label="Sobre",menu=menusobre)
Button(app,text="[1] - Cadastrar Devedor/Dívida ", command=cadastrar).place(x=10,y=100,width=200,height=40)
Button(app,text="[2] - Listar Devedores", command=leitura).place(x=10,y=150,width=200,height=40)
Button(app,text="[3] - Exibir Relatório detalhado", command=relatorio).place(x=10,y=200,width=200,height=40)
Button(app,text="[4] - Deletar Devedor", command=excluir).place(x=10,y=250,width=200,height=40)
Button(app,text="[5] - Sair do Programa", command=sair).place(x=10,y=300,width=200,height=40)



lb1 = Label(quadro, text="Bem Vindo!!", bg="#5f9ea0", font=("Arial",14),justify="center")
lb1.pack()
lb2 = Label(quadro2, text="", bg="#5f9ea0",font=("arial",11),justify="center")
lb2.place(x=0,y=5,width=285)
lb3 = Label(quadro, text="", bg="#5f9ea0",font=("arial",12))
lb3.place(x=1, y=50)



app.mainloop()